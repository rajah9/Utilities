"""
These classes opens and compares two Excel worksheets. (May be from the same file.).
It also reads Excel spreadsheets, reads rows or columns, and writes them to another spreadsheet.

-ExcelUtil does helpful conversions, such as convert_from_a1 which converts C4 to the row, col pair of 4,3.
 It is the parent of these child classes:
 - ExcelCompareUtil - Compares two files
 - ExcelRewriteUtil - Rewrites part of an Excel spreadsheet

To import this library, here's a handy import statement:
from ExcelUtil import ExcelCompareUtil

# To create the class, here's a handy instantiation statement:
eu = ExcelCompareUtil()
"""

import functools
import numbers
from collections import namedtuple
from copy import copy
from datetime import datetime
from math import fabs
from subprocess import CalledProcessError
from typing import List, Union, Tuple
from abc import abstractmethod
import numpy as np
import pandas as pd
import pdfplumber
from tabula import read_pdf, convert_into, errors

from CollectionUtil import CollectionUtil
from DateUtil import DateUtil
from FileUtil import FileUtil
from LogitUtil import logit
from PandasUtil import PandasUtil
from StringUtil import StringUtil, LineAccumulator
from Util import Util

_SUBPERIOD_DIVIDED = 'divided'
_SUBPERIOD_EQUAL = 'equal'

Strings = List[str]
Ints = List[int]
Floats = List[float]
Dataframes = List[pd.DataFrame]

_EPSILON = 1.0e-8
_SIG_CHARS = 12
ExcelCell = namedtuple('ExcelCell', 'col row')

"""
Interesting Python features:
* Uses a namedtuple for ExcelCell.
* Main class inherits from Util and has two subclasses.
* Several functions return a Tuple of two ints.
* Uses functools.lru_cache to cache up to 4 spreadsheets. Very helpful to keep the same thing from being read twice.
* In replace_col_with_scalar, uses two different methods to replace a masked array with a scalar.
* In replace_col_using_func, I pass in a function to be executed by it.
* In stats, provided an independent way to find slope, intercept, and r.
* In the init, set the display options to show head() better.
"""

class ExcelUtil(Util):
    def __init__(self):
        super().__init__()
        self.logger.info('starting ExcelUtil')
        self._pu = PandasUtil()
        self._su = StringUtil()
        self._wb = None

    def row_col_to_ExcelCell(self, row: int, col: int) -> ExcelCell:
        """
        Convert the given row and column to an ExcelCell.
        :param row:
        :param col:
        :return:
        """
        return ExcelCell(col=col, row=row)

    def ExcelCell_to_row_col(self, ec: ExcelCell) -> Tuple[int, int]:
        """
        Convert an ExcelCell to a corresponding ExcelCell.
        :param ec:
        :return:
        """
        return ec.row, ec.col

    def convert_from_A1(self, convert_me: str) -> Tuple[int, int]:
        row_col = StringUtil(convert_me)
        col = row_col.excel_col_to_int()
        row = row_col.digits_only_as_int()
        return row, col

    def ExcelCell_to_A1(self, ec: ExcelCell) -> str:
        """
        convert an ExcelCell to an Excel location
        :param ec: ExcelCell
        :return: str
        """
        return f'{self._su.int_to_excel_col(ec.col)}{ec.row}'

    def row_col_to_A1(self, row: int, col: int) -> str:
        """
        Convert the given row and column to an A1 format.
        :param row:
        :param col:
        :return: Cell in A1 (Excel) format
        """
        ec = self.row_col_to_ExcelCell(row, col)
        return self.ExcelCell_to_A1(ec)

    def convert_from_A1_to_cell(self, convert_me: str) -> ExcelCell:
        """
        Convert a string like A2 to an ExcelCell
        :param range: string like A2
        :return: ExcelCell like Excel(1,2)
        """
        row, col = self.convert_from_A1(convert_me)
        return self.row_col_to_ExcelCell(row=row, col=col)

    def convert_range_to_cells(self, range: str) -> Tuple[str, str]:
        """
        Convert a range like "a2:a9" to a pair of cells
        :param range: str like "a2:a9"
        :return: Tuple of strings on either side of the colon, like ("a2", "a9")
        """
        str_range = StringUtil(range)
        cell_range = str_range.split_on_delimiter(delim=':')
        if len(cell_range) == 2:
            return cell_range[0], cell_range[1]
        self.logger.warning(f'Should have exactly one cell before a colon and one after, but is: {range}. Returning (None, None)')
        return None, None

    def get_excel_rectangle_start_to(self, start: str, to: str) -> list:
        """
        Return a rectangle of ExcelCells.
        :param start: str containing a cell, like "A2"
        :param to: str containing a cell, like "A9"
        :return:
        """
        start_row, start_col = self.convert_from_A1(start)
        end_row, end_col = self.convert_from_A1(to)
        ans = [self.row_col_to_ExcelCell(col=col, row=row) for row in range(start_row, end_row + 1) for col in range(start_col, end_col + 1)]
        return ans

    def get_excel_rectangle(self, excel_range: str) -> list:
        """
        Return a list of ExcelCells
        :param excel_range: str like "a2:a9"
        :return: list of ExcelCells
        """
        start, to = self.convert_range_to_cells(excel_range)
        return self.get_excel_rectangle_start_to(start, to)

    def get_values(self, df: pd.DataFrame, rectangle: list) -> list:
        """
        Return the values described in rectangle.
        :param df:
        :param rectangle: a list, such as returned by get_excel_rectangle_start_to
        :return:
        """
        ans = []
        for cell in rectangle:
            val = df.iloc[cell.row - 2, cell.col - 1] # row is -2 because of -1 for header for 0-offset. col is -1 b/c of 0-offset
            ans.append(val)
        return ans

    @functools.lru_cache(maxsize=4)
    def load_spreadsheet(self, excelFileName: str = None, excelWorksheet: str = None) -> pd.DataFrame:
        df = self._pu.read_df_from_excel(excelFileName = excelFileName, excelWorksheet=excelWorksheet)
        return df

    def get_excel_filename_and_worksheet(self, excel_file_dict: dict) -> Tuple[str, str]:
        """
        Return the filename and worksheet in the given dictionary. If either is missing, return None for the missing field.
        :param excel_file_dict:
        :return:
        """
        filename, worksheet = None, None
        try:
            filename = excel_file_dict['filename']
        except KeyError:
            self.logger.warning(f'File name required in dictionary {excel_file_dict}, but it is missing.')
        try:
            worksheet = excel_file_dict['worksheet']
        except KeyError:
            self.logger.warning(f'Worksheet required in dictionary {excel_file_dict}, but it is missing.')
        return filename, worksheet

    def get_excel_filename_and_worksheet_and_range(self, excel_file_dict: dict) -> Tuple[str, str, str]:
        """
        Return the filename, worksheet and range in the given dictionary. If any is missing, return None for the missing field.
        :param excel_file_dict:
        :return:
        """
        filename, worksheet = self.get_excel_filename_and_worksheet(excel_file_dict)
        range = None
        try:
            range = excel_file_dict['range']
        except KeyError:
            self.logger.warning(f'Range required in dictionary {excel_file_dict}, but it is missing.')
        return filename, worksheet, range

    def _generate_excel_dict(self, excel_file_name: str, excel_worksheet: str, excel_range: str):
        excel_file_dict = {'filename': excel_file_name, 'worksheet': excel_worksheet, 'range': excel_range}
        return excel_file_dict

    def get_spreadsheet_values(self, excel_file_dict: dict) -> list:
        """
        Return the values specified by the efd.filename, efd.worksheet, and efd.range.
        If there is a (positive) efd.step=n, then keep the first value, step by n.
        :param excel_file_dict:
        :return:
        """
        fn, wks, range = self.get_excel_filename_and_worksheet_and_range(excel_file_dict)
        df = self.load_spreadsheet(excelFileName=fn, excelWorksheet=wks)
        area = self.get_excel_rectangle(excel_range=range)
        vals = self.get_values(df, area)
        excel_file_dict.setdefault('step', 1)
        my_step = excel_file_dict['step']
        return CollectionUtil.slice_list(my_list=vals, step=my_step)

    def get_scaling(self, excel_file_dict: dict) -> float:
        """
        Read the excel_file_dict for the 'scaling' field. If there's none, return 1.0
        :param excel_file_dict:
        :return:
        """
        try:
            scaling = excel_file_dict['scaling']
        except KeyError:
            self.logger.debug("no scaling found; using 1.0")
            scaling = 1.0
        return scaling

    def get_repeat(self, excel_file_dict: dict) -> Tuple[int, str] :
        """
        Return a tuple of a repeat factor and a string corresponding to the subperiods key.
        :param excel_file_dict: dict with 'repeat' key and 'subperiods' key.
        :return:
        """
        try:
            repeat = excel_file_dict['repeat']
        except KeyError:
            self.logger.debug("no repeat found; using 1")
            return 1, 'none'
        try:
            func_name = excel_file_dict['subperiod']
        except KeyError:
            self.logger.debug("No subperiod found; using equal")
            return repeat, _SUBPERIOD_EQUAL
        if func_name == _SUBPERIOD_EQUAL:
            return repeat, _SUBPERIOD_EQUAL
        elif func_name == _SUBPERIOD_DIVIDED:
            return repeat, _SUBPERIOD_DIVIDED
        else:
            self.logger.warning(f'Subperiod {func_name} not implemented yet.')
            return repeat, 'none'

    def get_epsilon(self, excel_file_dict: dict) -> float:
        """
        Read the excel_file_dict for the 'epsilon' field. If there's none, return 1.0
        :param excel_file_dict:
        :return:
        """
        try:
            return excel_file_dict['epsilon']
        except KeyError:
            self.logger.debug("no epsilon found; using class default")
        return None

    @abstractmethod
    def decode_named_range(self, excel_file_dict: dict) -> dict:
        pass

"""
To import this library, here's a handy import statement:
from ExcelUtil import ExcelCompareUtil

# Handy instantiation statement:
eu = ExcelCompareUtil()

"""
class ExcelCompareUtil(ExcelUtil):
    def __init__(self, epsilon: float = None):
        super().__init__()
        if epsilon:
            epsilon_str = f'passed-in value of {epsilon}'
            self._epsilon = epsilon
        else:
            epsilon_str = f'default of {_EPSILON}'
            self._epsilon = _EPSILON

        self.logger.info('starting ExcelCompareUtil with ' + epsilon_str)
        self._compare_log = LineAccumulator()

    # Getters and setters for epsilon
    @property
    def epsilon(self):
        return self._epsilon

    # Setter for epsilon.
    @epsilon.setter
    def epsilon(self, eps: float):
        self._epsilon = eps

    # Getter for compare_log.
    @property
    def compare_log(self) -> Strings:
        return self._compare_log.contents # Not the compare_log (a LineAccumulator) but its contents (a list of strings)

    def add_log_line(self, line: str):
        self._compare_log.add_line(line)

    def close_numbers(self, list1: Floats, list2: Floats, scaling: float = 1.0, epsilon: float = None) -> bool:
        """
        test list1 against list2. return True if all the numbers are within epsilon.
        If not set, use the default epsilon for the class.
        :param list1:
        :param list2:
        :param scaling:
        :param epsilon:
        :return:
        """
        eps = epsilon or self._epsilon
        ans = True
        for el1, el2 in zip(list1, list2):
            if fabs(el1 - scaling * el2) > eps:
                ans = False
                self.logger.warning(f'mismatch: {el1} not equal to {el2} * scale {scaling}; difference of {fabs(el1 - el2 * scaling)}')
        return ans

    def identical_ints(self, list1: Ints, list2: Ints, scaling: float = 1.0):
        """
        determine if the list of integers is identical.
        :param list1:
        :param list2:
        :param scaling:
        :return:
        """
        ans = True
        for el1, el2 in zip(list1, list2):
            if el1 != scaling * el2:
                ans = False
                self.logger.warning(f'mismatch: {el1} not equal to {el2} * scale {scaling}; difference of {abs(el1 - el2 * scaling)}')
        return ans

    def identical_strings(self, list1: Strings, list2: Strings, significant_characters: int = None) -> bool:
        """
        Compare the strings in the lists, up to the significant_character.
        If significant_character is None, then compare the whole string.
        Return True if all of the values are equal to each other.
        Allow for the floats to be within epsilon and still be considered identical.
        :param list1: List of str
        :param list2: list of str
        :param significant_characters: None to compare each el in totality, or an int n to compare the first n chars.
        :return: True IFF each element matches.
        """
        ans = True
        for el1, el2 in zip(list1, list2):
            if significant_characters:
                if el1[:significant_characters] != el2[:significant_characters]:
                    ans = False
                    self.logger.warning(f'mismatch: {el1} not equal to {el2} within the first {significant_characters} characters.')
            else:
                if el1 != el2:
                    ans = False
                    self.logger.warning(f'mismatch: {el1} not equal to {el2}.')

        return ans

    def identical(self, list1:  Union[Ints, Floats, Strings], list2:  Union[Ints, Floats, Strings], scaling: Union[float, int] = 1, epsilon: float = None) -> bool:
        """
        Determine if these strings, ints, or floats are identical on an element-by-element comparison.
        :param list1:
        :param list2:
        :param scaling:
        :param epsilon:
        :return:
        """
        if len(list1) != len(list2):
            self.logger.warning(f'Lists are different sizes. {len(list1)} not equal to {len(list2)}')
            return False

        list1_el = list1[0]
        list2_el = list2[0]

        # Both strings
        if isinstance(list1_el, str) and isinstance(list2_el, str):
            return self.identical_strings(list1, list2, significant_characters=None)

        # Both ints
        if isinstance(list1_el, (int, np.int64)) and isinstance(list2_el, (int, np.int64)):
            return self.identical_ints(list1, list2, scaling)

        # If comparing ints to floats, then coerce one vector to floats.
        if isinstance(list1_el, (int, np.int64)):
            list1 = [float(x) for x in list1]
            list1_el = list1[0]
        if isinstance(list2_el, (int, np.int64)):
            list2 = [float(x) for x in list2]
            list2_el = list2[0]

        # Both floats
        if isinstance(list1_el, float) and isinstance(list2_el, float):
            return self.close_numbers(list1, list2, scaling, epsilon)

        # raise a warning about types.
        self.logger.warning(f'first element of list1 is {list1_el}, but first element of list2 is {list2_el}. They cannot be compared. Returning False. ')
        return False

    def verify(self, file1: dict, file2: dict) -> bool:
        """
        Verify that the values in rectangles of input_file_dict and output_file_dict are identical.
        :param file1: dict with keys 'filename', 'worksheet', and 'range'
        :param file2:
        :return:
        """
        vals1 = self.get_spreadsheet_values(excel_file_dict=file1)
        self.logger.debug(f'1: \n{vals1}')
        vals2 = self.get_spreadsheet_values(excel_file_dict=file2)

        scaling = self.get_scaling(file2)

        self.logger.debug(f'2:\n{vals2}')
        eps = self.get_epsilon(file2)
        file1_fn, file1_wks, file1_range = self.get_excel_filename_and_worksheet_and_range(file1)
        file2_fn, file2_wks, file2_range = self.get_excel_filename_and_worksheet_and_range(file2)
        self.add_log_line(f"{0:^23} file {1:100}, worksheet {2:15}, and range {3}".format("Starting comparison of", file1_fn, file1_wks, file1_range))
        self.add_log_line(f"{0:^23} file {1:100}, worksheet {2:15}, and range {3}".format("Starting comparison of", file2_fn, file2_wks, file2_range))
        ans = self.identical(vals1, vals2, scaling=scaling, epsilon=eps)
        report = 'identical' if ans else 'DIFFERENT'
        self.logger.info(f'lists are {report}')
        self.add_log_line(f"{0:^23}:{1}".format("RESULTS", report))
        return ans

    def compare_list_els_against_scalar(self, vals: list, compare_me: Union[float, str]) -> bool:
        """
        Compare the elements in the list against the scalar and return true if they are all equal.
        :param list1: list like [13, 13, 13, 13] or ['Figaro', 'Figaro', 'Figaro']
        :param compare_me: scalar like 13 or 'Figaro'
        :return: True if all elements in list1 equal compare_me
        """
        scalars = [compare_me] * len(vals)
        return self.identical(vals, scalars)

    def compare_to_scalar(self, file_dict: dict, compare_me: Union[float, str]) -> bool:
        """
        Compare the spreadsheet values and range to the scalar compare_me.
        Return true if all of the values are equal to the scalar.
        :param file_dict: dict with 'filename', 'worksheet', and 'range'
        :param compare_me: scalar to compare against
        :return: True if all elements in the given range are equal to compare_me.
        """
        vals = self.get_spreadsheet_values(excel_file_dict=file_dict)
        return self.compare_list_els_against_scalar(vals, compare_me)

"""
Following routines are for reading and writing Excel files.
They use openpyxl.
See generates_spreadsheets.py as an example.
"""
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelRewriteUtil(ExcelUtil):
    def __init__(self):
        super().__init__()
        self._su = StringUtil()
        self._fu = FileUtil()
        self._wb = None

    @functools.lru_cache(maxsize=2)
    def init_workbook_to_read(self, filename: str) -> Workbook:
        self._wb = load_workbook(filename=filename)
        return self._wb


    def init_workbook_to_write(self) -> Workbook:
        self._wb = Workbook()

    @logit(showArgs=True, showRetVal=False)
    def save_workbook(self, filename: str):
        self._wb.save(filename=filename)

    def worksheet_names(self) -> list:
        """
        Return a list of the worksheet names.
        :return:
        """
        return self._wb.sheetnames

    def index_of_worksheet_name(self, find_this_worksheet: str) -> int:
        """
        Find the index of the given worksheet by name.
        :param find_this_worksheet: name of worksheet, like 'Sheet'
        :return: 0-based offset of the worksheet, or raise a ValueError if not found.
        """
        ws_names = self.worksheet_names()
        try:
            return ws_names.index(find_this_worksheet)
        except ValueError as e:
            self.logger.warning(f'Unable to find the worksheet named {find_this_worksheet}')
            raise e

    def set_active(self, make_active: str) -> bool:
        """
        Set the given sheet to active.
        :param make_active: name of the sheet to make active.
        :return: True if successfully made active.
        """
        try:
            index = self.index_of_worksheet_name(make_active)
            self._wb.active = index
            return True
        except ValueError:
            return False

    def get_cell(self, ws: Worksheet, cell_loc: str = 'A1', want_value: bool = True) -> Union[int, float, str]:
        """
        Return the cell at the given ExcelCell.
        :param cell_loc: cell location using Excel notation
        :return: value at that cell
        """
        if want_value:
            return ws[cell_loc].value
        return ws[cell_loc]

    def get_cells(self, ws: Worksheet, excel_range: str = 'A1:C3') -> list:
        """
        Return the cells in the given range.
        :param ws:
        :param excel_range:
        :return:
        """
        start, to = self.convert_range_to_cells(excel_range) # start like 'A1'; to like 'C3'
        cells = self.get_excel_rectangle_start_to(start, to) # list of cells
        ans = []
        for cell in cells:
            cell_loc = self.ExcelCell_to_A1(cell)
            ans.append(self.get_cell(ws, cell_loc))
        return ans

    def _scaled(self, output_file_dict: dict, vals: list) -> list:
        """
        Return a scaled version of the vals list, multiplying by the given factor (or 1.0).
        :param output_file_dict: dict with (optional) key 'scaling'
        :param vals:
        :return:
        """
        if isinstance(vals[0], str):
            return vals # First item in list a str, so no scaling needed.

        scaling = self.get_scaling(output_file_dict)
        scaled_vals = [v * scaling for v in vals]
        return scaled_vals

    def _repeated(self, output_file_dict: dict, vals: list) -> list:
        """
        Repeat each element. If the 'repeat' key is 3, then also look at the subperiods key.
          subperiods   action
          equal        repeat each element 3 times. (default)
          divided      divide each element by 3
        :param output_file_dict: dict with an (optional) key 'repeat' and an (optional) key subperiods.
        :param vals:
        :return:
        """
        repeat, subperiod = self.get_repeat(output_file_dict)
        if repeat == 1:
            return vals
        if subperiod == _SUBPERIOD_EQUAL:
            return self._equal_subperiod(vals, repeat)
        elif subperiod == _SUBPERIOD_DIVIDED:
            return self._divided_subperiod(vals, repeat)
        else:
            return self._no_subperiod()

    def _no_subperiod(self):
        """
        raise an error.
        :return:
        """
        raise NotImplementedError('no subperiod defined')

    def _equal_subperiod(self, vals: list, repeat: int):
        """
        Repeat each el in vals repeat times
        :param vals: list, like [2, 4, 6, 8]
        :param repeat: int, like 2
        :return: list, like [2, 2, 4, 4, 6, 6, 8, 8]
        """
        ans = CollectionUtil.repeat_elements_n_times(vals, repeat)
        return ans

    def _divided_subperiod(self, vals: list, repeat: int):
        """
        Repeat each el in vals repeat times, dividing each el by repeat.
        :param vals: list, like [2, 4, 6, 8]
        :param repeat: int, like 2
        :return: list, like [1, 1, 2, 2, 3, 3, 4, 4]
        """
        ans = CollectionUtil.repeat_elements_n_times(vals, repeat)
        return [el / repeat for el in ans]


    def load_and_write(self, input_file_dict: dict, output_file_dict: dict, do_save: bool = True) -> bool:
        """
        Load the Excel file and worksheet in input_file_dict and write it (with scaling) to output_file_dict
        :param input_file_dict: dict with keys 'filename', 'worksheet' and 'range' or 'ranges'
        :param output_file_dict: dict with keys 'filename', 'worksheet' and 'range' or 'ranges'
        :return:
        """
        vals1 = self.get_spreadsheet_values(excel_file_dict=input_file_dict)
        self.logger.debug(f'1: \n{vals1}')
        vals2 = self._repeated(output_file_dict, vals1)
        scaled_vals = self._scaled(output_file_dict, vals2)
        # We have either a single range or many ranges to copy to.
        try:
            rectangles = output_file_dict['ranges']
        except KeyError:
            range_rectangle = output_file_dict['range']
            rectangles = [range_rectangle]

        self.write_range_to_worksheet(excel_worksheet=output_file_dict['worksheet'], ranges=rectangles, vals=scaled_vals)
        if do_save:
            self.save_workbook(output_file_dict['filename'])
        return True

    @functools.lru_cache(maxsize=2)
    def init_workbook(self, filename: str) -> Workbook:
        self._wb = load_workbook(filename=filename)
        return self._wb

    @logit()
    def rewrite_worksheet(self, excel_filename: str, excel_worksheet: str, ranges: list, vals: list):
        """
        Read in the given filename and worksheet (from the file dictionary).
        Write the values to the given range, preserving formatting.

        :param excel_filename:
        :param excel_worksheet:
        :param ranges:
        :param vals:
        :return:
        """
        wb = self.init_workbook_to_read(excel_filename)
        ws = wb[excel_worksheet]
        # Ensure that the sum of the area of the ranges equals the number of variables
        local_vals = copy(vals)
        area = 0
        for rectangle in ranges:
            area += len(self.get_excel_rectangle(rectangle))  # Will return a list of ExcelCells
        if area != len(vals):
            self.logger.warning(f'mismatch between source length (={len(vals)}) and target length (={area}) defined by cell ranges {ranges}')
        for rectangle in ranges:
            excel_rect = self.get_excel_rectangle(rectangle)
            start_cell, end_cell = excel_rect[0], excel_rect[-1]
            self.logger.debug(f'starting cell: {start_cell} / ending cell: {end_cell}')

            for row in range(start_cell.row, end_cell.row + 1):
                for col in range (start_cell.col, end_cell.col + 1):
                    val = local_vals.pop(0)
                    self.logger.debug(f'cell at row {row} / col {col} is {ws.cell(row=row, column=col, value=val).value}')
        return

    def write_range_to_worksheet(self, excel_worksheet: str, ranges: list, vals: list, init_workbook: bool = True):
        """
        Write the values to the given range, preserving formatting.
        Refactor of rewrite_worksheet.
        :param file: dictionary containing 'filename', 'worksheet', and 'range' keys.
        :param vals: list of values to be written. len(vals) should be the same as the range.
        :return:
        """
        if init_workbook:
            wb = self.init_workbook_to_write()
        ws = self._create_worksheet(excel_worksheet)
        # Ensure that the sum of the area of the ranges equals the number of variables
        local_vals = copy(vals)
        area = 0
        for rectangle in ranges:
            area += len(self.get_excel_rectangle(rectangle))  # Will return a list of ExcelCells
        if area != len(vals):
            self.logger.warning(f'mismatch between source length (={len(vals)}) and target length (={area}) defined by cell ranges {ranges}')
        for rectangle in ranges:
            excel_rect = self.get_excel_rectangle(rectangle)
            start_cell, end_cell = excel_rect[0], excel_rect[-1]
            self.logger.debug(f'starting cell: {start_cell} / ending cell: {end_cell}')

            for row in range(start_cell.row, end_cell.row + 1):
                for col in range (start_cell.col, end_cell.col + 1):
                    val = local_vals.pop(0)
                    self.logger.debug(f'cell at row {row} / col {col} is {ws.cell(row=row, column=col, value=val).value}')
        return


    def write_df_to_excel(self, df: pd.DataFrame, excelFileName:str, excelWorksheet:str="No title", attempt_formatting:bool=False, write_header=False, write_index=False):
        """
        Write the given dataframe to Excel. Attempt to format percents and numbers as percents and numbers, if requested.
        Code adapted from https://openpyxl.readthedocs.io/en/stable/pandas.html .
        :param df:
        :param excelFileName:
        :param attempt_formatting:
        :return:
        """
        self.write_df_to_ws(df=df, excelWorksheet=excelWorksheet, attempt_formatting=attempt_formatting, write_header=write_header, write_index=write_index)
        self.save_workbook(excelFileName)

    def write_df_to_new_ws(self, df: pd.DataFrame, excelWorksheet: str = "No title", attempt_formatting: bool = False, write_header: bool = False, write_index: bool = False) -> Worksheet:
        """
        Initialize the workbook. Then write the given dataframe to an excel worksheet.
        Attempt to format percents and numbers as percents and numbers, if requested.
        Code adapted from https://openpyxl.readthedocs.io/en/stable/pandas.html .

        :param df:
        :param excelWorksheet: string of worksheet name
        :param write_header:
        :param write_index:
        :return:
        """
        self.init_workbook_to_write()
        return self.write_df_to_ws(df=df, excelWorksheet=excelWorksheet, attempt_formatting=attempt_formatting, write_header=write_header, write_index=write_index)

    def write_df_to_ws(self, df: pd.DataFrame, excelWorksheet: str = "No title", attempt_formatting:bool = False, write_header = False, write_index = False, date_format: str = 'MM/DD/YYYY') -> Worksheet:
        """
        Refactor of write_df_to_new_ws, which does not init the workbook.
        Write the given dataframe to an excel worksheet. Attempt to format strings ending in % and strings as numbers as percents and numbers, if requested.
        Code adapted from https://openpyxl.readthedocs.io/en/stable/pandas.html .

        :param df: dataframe to write
        :param excelWorksheet:
        :param attempt_formatting: True means attempt to format strings as percents or numbers
        :param write_header: True if you'd like to write the column names
        :param write_index: True if you'd like to write the index
        :param date_format: Date (or time) format, such as 'dd/mm/yy' or 'YY MMM DD'
               More styles may be found at https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.htm.
        :return:
        """
        ws = self._create_worksheet(excelWorksheet)
        formatting = {'Normal': 'Normal', 'Percent': '#0.00%', 'Comma': '#,##0.00'}
        # Write the whole dataframe
        for row in dataframe_to_rows(df, index=write_index, header=write_header):
            ws.append(row)
        if attempt_formatting:  # apply formatting if requested.
            for row in ws.iter_rows(min_row=ws.min_row, min_col=ws.min_column, max_row=ws.max_row,
                                    max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, str):
                        c = self._su.convert_string_append_type(cell.value)
                        cell.value = c.value
                        if (c.cellType == 'Comma') or (c.cellType == 'Percent'):
                            cell.number_format = formatting[c.cellType]
                    elif isinstance(cell.value, datetime):
                        cell.number_format = date_format
                    else:
                        pass
        return ws

    def _create_worksheet(self, excelWorksheet: str) -> Worksheet:
        """
        Create the worksheet. If it exists, return the existing worksheet.
        :param excelWorksheet:
        :return:
        """
        if excelWorksheet in self.worksheet_names():
            self.logger.warning(f'Already have sheet name {excelWorksheet}!')
            ws = self._wb[excelWorksheet]
        else:
            self.logger.debug(f'Creating new worksheet: {excelWorksheet}')
            ws = self._wb.create_sheet(title=excelWorksheet)
        return ws

    def copy_ws_to_ws(self, ws_source: Worksheet, ws_source_name: str, ws_target_name: str = None, range: str = None):
        """
        Copy from the given worksheet to the currently active workbook.
        Based on the accepted answer at https://stackoverflow.com/questions/50208440/openpyxl-copy-from-one-workbook-to-another.
        :param ws_source: Source worksheet
        :return: None
        """
        self.init_workbook_to_write()
        new_ws_name = ws_target_name or ws_source_name
        ws_dest = self._wb.create_sheet(new_ws_name)
        if not range:
            start_row = ws_source.min_row # Usually 1
            start_col = ws_source.min_column # usually 1
            end_row = ws_source.max_row
            end_col = ws_source.max_column
        else:
            raise NotImplementedError('must set range to None. Partial ranges not implemented yet.')
        source_range = self.row_col_to_A1(row=start_row, col=start_col) + ":" + self.row_col_to_A1(row=end_row, col=end_col)
        for row in ws_source[source_range]:
            for cell in row:
                ws_dest[cell.coordinate].value = ws_source[cell.coordinate].value
                # Following code adopted from https://stackoverflow.com/a/34838233/509840
                if cell.has_style:
                    ws_dest[cell.coordinate].font = copy(cell.font)
                    ws_dest[cell.coordinate].border = copy(cell.border)
                    ws_dest[cell.coordinate].fill = copy(cell.fill)
                    ws_dest[cell.coordinate].number_format = copy(cell.number_format)
                    ws_dest[cell.coordinate].protection = copy(cell.protection)
                    ws_dest[cell.coordinate].alignment = copy(cell.alignment)

    def copy_spreadsheet_to_ws(self, sourceFileName: str, sourceWorksheet: str, destWorksheet: str = None,
                               header: int = 0, attempt_formatting: bool = False, write_header: bool = False,
                               write_index: bool = False, create_new_ws: bool = True, input_range=None,
                               output_range=None) -> Worksheet:
        """
        Read the sourceWorksheet into a dataframe and write it to the destWorksheet.
        :param sourceFileName:
        :param sourceWorksheet:
        :param destWorksheet: Title of the destination worksheet (or copy the sourceWorksheet if None)
        :param header:
        :param attempt_formatting:
        :param write_header:
        :param write_index:
        :param create_new_ws: set True to create a new worksheet. False uses the existing.
        :param input_range: string like A9:A15 (8 cells down)
        :param output_range: string like A12:H12 (8 cells across)
        :return: copied worksheet
        """
        # 1. Read the existing sourceWorksheet.
        df = self._pu.read_df_from_excel(excelFileName=sourceFileName, excelWorksheet=sourceWorksheet, header=header)
        if len(df):
            self.logger.debug(f'Read in {len(df)} records from file {sourceFileName} and worksheet {sourceWorksheet}.')
            worksheet_name = destWorksheet or sourceWorksheet
            # 2. Create a ws_source based on the df.
            if create_new_ws:
                ws = self.write_df_to_new_ws(df=df, excelWorksheet=worksheet_name, attempt_formatting=attempt_formatting, write_header=write_header, write_index=write_index)
                return ws
            else:
                if not input_range:
                    ws = self.write_df_to_ws(df=df, excelWorksheet=worksheet_name, attempt_formatting=attempt_formatting, write_header=write_header, write_index=write_index)
                    return ws
                else:
                    file_dict = self._generate_excel_dict(excel_file_name=sourceFileName, excel_worksheet=worksheet_name, excel_range=input_range)
                    vals = self.get_spreadsheet_values(excel_file_dict=file_dict)
                    self.write_range_to_worksheet(excel_worksheet=worksheet_name, ranges=[output_range], vals=vals, init_workbook=False)
            return

        self.logger.warning(f'Read in no records from file {sourceFileName} and worksheet {sourceWorksheet}. Returning an empty ws_source')
        return None


    def init_template(self, template_excel_file_name: str, template_excel_worksheet: str, output_excel_file_name: str,
                      output_excel_worksheet: str = None, excel_range: str = None) -> bool:
        """
        Read from the given template_excel_file_name and template_excel_worksheet.
        Copy the template info into the output file and output worksheet.
        Apply formatting from the template into the output worksheet.

        :param template_excel_file_name:
        :param template_excel_worksheet:
        :param output_excel_file_name:
        :param output_excel_worksheet:
        :param excel_range: (optional) excel range, like 'A1:G4'. If omitted, use the whole template.
        :return: True if the worksheet is set up.
        """
        if not self._fu.file_exists(qualifiedPath=template_excel_file_name):
            self.logger.warning(f'Could not find file {template_excel_file_name}!')
            return False
        if not output_excel_worksheet:
            output_excel_worksheet = template_excel_worksheet

        self.init_workbook_to_read(filename=template_excel_file_name)
        template_ws = self._wb[template_excel_worksheet]
        self.copy_ws_to_ws(ws_source=template_ws, ws_source_name=template_excel_worksheet, ws_target_name=output_excel_worksheet)
        self.save_workbook(filename=output_excel_file_name)

        # Following fails.
#        wc = WorksheetCopy(template_ws, output_ws) # Create the copy object, providing source and target ws_source
#        wc.copy_worksheet() # copy the requested worksheet
        return True


    def read_template(self, output_dict: dict, template_dict: dict) -> Tuple[str, str]:
        """
        Read the template file and create an active worksheet.
        Modified 26Jan21.
        :param output_dict:
        :param template_dict:
        :return:
        """
        template_fn, template_wks, template_range = self.get_excel_filename_and_worksheet_and_range(template_dict)
        output_fn, output_wks, output_range = self.get_excel_filename_and_worksheet_and_range(output_dict)
        self.init_workbook_to_write()
        new_index = 0  # adding 26Jan21
        self._wb.create_sheet(output_wks, index=new_index)  # adding 26Jan21
        self._wb.active = new_index  # adding 26Jan21
        self.copy_spreadsheet_to_ws(sourceFileName=template_fn, sourceWorksheet=template_wks,
                                         destWorksheet=output_wks, create_new_ws=True)
        return output_fn, output_wks


    def read_template_and_nodes(self, template_dict: dict, output_dict: dict, dicts: list):
        """
        Read the template file and create an active worksheet. Read the filenames and worksheets and ranges in dicts
        and write the ranges to the active worksheet. Finally write the worksheet to the filename in the output_dict.
        Can be called like this:
                    in_file_dict = {
                'filename': self.formatting_spreadsheet_name,
                'worksheet': self.worksheet_name,
                'range': 'B3:B6'
            }

            # output file is first.xlsx.
            out_file_dict = {
                'filename': self.parent_spreadsheet_name,
                'worksheet': self.worksheet_name,
                'range': 'A1:A4'
            }
            # Node file is node.xlsx.
            node_file_dict = {
                'filename': self.node_spreadsheet_name,
                'worksheet': self.worksheet_name,
                'range': 'b2:b6',
                'outputrange': 'a7:e7'
            }
            dict_list = [node_file_dict]
            self._rwu.read_template_and_nodes(template_dict=in_file_dict, output_dict=out_file_dict, dicts=dict_list)

        :param template_dict: dict with 'filename' and 'worksheet' keys.
        :param output_dict:  dict with 'filename' and 'worksheet' keys.
        :param dicts: dict with  'filename' and 'worksheet' and 'range' or 'namedrange' keys.
        :return:
        """
        output_fn, output_wks = self.read_template(output_dict, template_dict)
        # Now loop through the list of dicts.
        for file_dict in dicts:
            file_fn, file_wks, file_range = self.get_excel_filename_and_worksheet_and_range(file_dict)
            self.logger.debug(f'processing spreadsheet {file_fn} in worksheet {file_wks} with range {file_range}')
            try:
                out_range = file_dict['outputrange']
                self.copy_spreadsheet_to_ws(sourceFileName=file_fn, sourceWorksheet=file_wks, destWorksheet=output_wks,
                                            create_new_ws=False, input_range=file_range, output_range=out_range)
            except KeyError:
                self.logger.warning(f'The node dictionary {file_dict} is missing the key <outputrange>. Returning.')
                return
        self.save_workbook(filename=output_fn)
        return

    def stream_df_to_ws(self, df: pd.DataFrame, worksheet_name: str):
        """
        Stream the given df to the end of the active worksheet.
        Code adopted from the openpyxl documentation: https://openpyxl.readthedocs.io/en/stable/pandas.html
        :param df:
        :return:
        """
        if not self.set_active(make_active=worksheet_name):
            self.logger.warning(f'Could not find spreadsheet named {worksheet_name}! Not writing further.')

        ws = self._wb.active

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

class PdfToExcelUtil(ExcelUtil):
    def __init__(self):
        super().__init__()
        self.logger.info('starting PdfToExcelUtil')
        self._cu = CollectionUtil()

    def read_pdf_table(self, pdf_filename: str, pages: Union[Ints, str] = 'all', make_NaN_blank: bool = True,
                       read_many_tables_per_page=False) -> Dataframes:
        pass

    def read_tiled_pdf_tables(self, pdf_filename: str, rows: int, cols: int, pages: Union[Ints, str] = 'all',
                              tables_to_tile: Ints = None, tile_by_rows: bool = True, read_many_tables_per_page=False,
                              make_NaN_blank: bool = True) -> Dataframes:
        """
        This reads the tables on several pages into a single table.
        :param pdf_filename:
        :param rows: How many rows across
        :param cols: How many columns down
        :param pages: Default: 'all'. Could be a str like "1,2,3" or a list like [1,2,3]. ONE-offset.
        :param tables_to_tile: ONE-offset tables to be read in.
        :param tile_by_rows: True if sequential pages comprise rows, False if they comprise columns.
        :param read_many_tables_per_page: False to read one table per page.
        :param make_NaN_blank: True to make NaN values blank.
        :return:
        """
        dfs = self.read_pdf_table(pdf_filename, pages=pages, make_NaN_blank=make_NaN_blank,
                                  read_many_tables_per_page=read_many_tables_per_page)
        expected_table_count = len(tables_to_tile)
        are_tables_ok = True
        if expected_table_count != len(dfs):
            self.logger.warning(f'There are {len(tables_to_tile)} tables to read, but read in {len(dfs)}')
            if len(dfs) < expected_table_count:
                self.logger.warning(f'Returning no tables')
                are_tables_ok = False
                return None
            else:
                self.logger.warning(f'Will use the first {expected_table_count} tables.')
        zero_offset_tables_to_tile = [t - 1 for t in tables_to_tile]
        for i in zero_offset_tables_to_tile:
            if len(dfs[i]):
                self.logger.debug(f'Table {i} has {len(dfs[i])} records.')
            else:
                self.logger.warning(f'Table {i} is empty.')
                are_tables_ok = False

        if not are_tables_ok:
            self.logger.warning(f'Requested tables {tables_to_tile} but one or more were empty.')

        table_layout = self._cu.layout(rows, cols, row_dominant=tile_by_rows, tiling_order=zero_offset_tables_to_tile)

        # Assemble columns by rows
        row_dfs = []
        for i, row in enumerate(table_layout):
            self.logger.debug(f'row {i} contains: {row}')
            this_row_df = self._pu.join_dfs_by_column([dfs[x] for x in row])
            row_dfs.append(this_row_df)

        # Assemble the rows into one big DF
        big_df = self._pu.join_dfs_by_column(row_dfs)
        return big_df

    def summarize_pdf_tables(self, pdf_filename: str, pages: Union[Ints, str] = 'all', make_NaN_blank: bool = True,
                        read_many_tables_per_page:bool=False, how_many_summarized: int = 10) -> Strings:
        """
        Summarize the tables on the given pages, and return the first how_many_summarized lines of each table.
        :param pdf_filename: Filename to scan
        :param pages: Default: 'all'. Could be a str like "1,2,3" or a list like [1,2,3].
        :param make_NaN_blank: True to make NaN values blank (default). False to leave as NaN.
        :param read_many_tables_per_page:
        :return:
        """
        ans = LineAccumulator()
        dfs = self.read_pdf_table(pdf_filename, pages, make_NaN_blank, read_many_tables_per_page)
        for i, df in enumerate(dfs):
            header = self._su.fill_string(my_str=f'Table {i} contains {len(df)} records.')
            ans.add_line(header)
            if len(df) > 0:
                ans.add_df(df, how_many_rows=how_many_summarized)
        return ans.contents

class PdfToExcelUtilTabula(PdfToExcelUtil):
    def __init__(self):
        super().__init__()
        self.logger.info('starting PdfToExcelUtilTabula')

    def read_pdf_table(self, pdf_filename: str, pages: Union[Ints, str] = 'all', make_NaN_blank: bool = True,
                       read_many_tables_per_page=False) -> Dataframes:
        """
        Using tabula-py, read the designated pages from the PDF.
        :param pdf_filename: full path and filename to PDF file. May be a URL.
        :param pages: page to read in (defaults to 'all')
        :param read_many_tables_per_page: False to read one table per page.
        :return: a list of dataframes, where each el is a table
        """
        self.logger.debug(f'About to read from {pdf_filename} using tabula-py. Pages are 0-offset: {pages}')
#        dfs = tabula.read_pdf(pdf_filename, pages=pages, multiple_tables=read_many_tables_per_page)
        # tabula_build_opts = tabula.io.build_options(pages=pages, stream=True)
        dfs = read_pdf(pdf_filename, guess=False, pages=pages, stream=True, multiple_tables=read_many_tables_per_page, pandas_options={'header': None}) # problems with multiple_tables=read_many_tables_per_page)
        self.logger.debug(f'Read in {len(dfs)} tables.')
        if not make_NaN_blank:
            return dfs
        blanked_dfs = [df.replace(np.nan, '', regex=True) for df in dfs]
        return blanked_dfs

    def read_pdf_write_csv(self, pdf_filename: str, csv_filename: str, pages: Union[Ints, str] = 'all') -> bool:
        """
        Use tabula's convert_into routine to read a PDF and process into CSVs.
        Exceptions are from https://tabula-py.readthedocs.io/en/latest/tabula.html .
        :param pdf_filename:
        :param csv_filename:
        :param pages:
        :return:
        """
        try:
            convert_into(pdf_filename, csv_filename, output_format="csv", pages=pages)
            self.logger.info(f'Successfully wrote to {csv_filename}.')
            return True
        except FileNotFoundError:
            self.logger.warning(f"Can't find the input file {pdf_filename}")
            return False
        except ValueError:
            self.logger.warning(f'downloaded input file {pdf_filename} is of size 0')
            return False
        except errors.JavaNotFoundError:
            self.logger.warning(f'Java was not found')
            return False
        except CalledProcessError:
            self.logger.warning('tabula-java execution failed')
            return False

class PdfToExcelUtilPdfPlumber(PdfToExcelUtil):
    def __init__(self):
        super().__init__()
        self.logger.info('starting PdfToExcelUtilPdfPlumber')
        # Original table settings from https://github.com/jsvine/pdfplumber#extracting-tables
        self.table_settings = {"vertical_strategy": "lines", "horizontal_strategy": "lines",
                               "explicit_vertical_lines": [], "explicit_horizontal_lines": [], "snap_tolerance": 3,
                               "join_tolerance": 3, "edge_min_length": 3, "min_words_vertical": 3,
                               "min_words_horizontal": 1, "keep_blank_chars": False, "text_tolerance": 3,
                               "text_x_tolerance": None, "text_y_tolerance": None, "intersection_tolerance": 3,
                               "intersection_x_tolerance": None, "intersection_y_tolerance": None}
        self.table_settings["vertical_strategy"] = "text"
        self.table_settings["horizontal_strategy"] = "text"
        self.table_settings["text_tolerance"] = 9
        self.table_settings["text_x_tolerance"] = 9
        self.table_settings["text_y_tolerance"] = 5

    def read_pdf_table(self, pdf_filename: str, pages: Union[Ints, str] = 'all', make_NaN_blank: bool = True,
                       read_many_tables_per_page=False) -> Dataframes:
        """
        Using pdfplumber, read the designated pages from the PDF.
        :param pdf_filename: full path and filename to PDF file. May be a URL.
        :param pages: page to read in (defaults to 'all')
        :param read_many_tables_per_page: False to read one table per page.
        :return: a list of dataframes, where each el is a table
        """
        self.logger.debug(f'About to read from {pdf_filename} using PdfPlumber. Pages are 1-offset: {pages}')
        ans = []
        with pdfplumber.open(pdf_filename) as pdf:
            scanned_pages = pdf.pages
            for i, pg in enumerate(scanned_pages):
                tbl = scanned_pages[i].extract_tables(table_settings=self.table_settings)
                df = self._pu.convert_list_to_dataframe(tbl[0])
                pg_str = f'pg{i}col'
                self._pu.replace_col_names_by_pattern(df, prefix=pg_str, is_in_place=True)
                ans.append(df)
                self.logger.info(f'Table i {i} has {len(df)} records')
        return ans

    def summarize_pdf_tables(self, pdf_filename: str, pages: Union[Ints, str] = 'all', make_NaN_blank: bool = True,
                        read_many_tables_per_page:bool=False, how_many_summarized: int = 10) -> Strings:
        """
        Summarize the PDF tables in a PdfPlumber kinda way.
        This code adopted from: https://stackoverflow.com/questions/55939921/use-pdfplumber-to-find-text-in-pdf-return-page-number-then-return-table
        :param pdf_filename:
        :param pages:
        :param make_NaN_blank:
        :param read_many_tables_per_page:
        :param how_many_summarized:
        :return:
        """
        ans = LineAccumulator()
        with pdfplumber.open(pdf_filename) as pdf:
            scanned_pages = pdf.pages
            for i, pg in enumerate(scanned_pages):
                tbl = scanned_pages[i].extract_tables(table_settings=self.table_settings)
                rec_count = len(tbl[0])
                lines = self.clean_tables(tbl, how_many_summarized)
                header = self._su.fill_string(my_str=f'Table {i} contains {rec_count} records.')
                ans.add_line(header)
                ans.add_lines(lines)
                print(f'{i} --- {tbl}')
        return ans.contents

    def clean_tables(self, tables: list, how_many_summarized: int = 10) -> Strings:
        """
        tables is a nested list. Clean them up in a list of strings.
        :param how_many_summarized:
        :param tables: [ [ ['one column'] ['another column'] ]]
        :return:
        """
        ans = LineAccumulator()
        tbl = tables[0]
        for lst in tbl:
            if ans.contents_len() >= how_many_summarized:
                break
            null_to_spc = self._cu.replace_elements_in_list(lst, '', ' ')
            line = ''.join(null_to_spc)
            ans.add_line(line)
        return ans.contents


"""
DfHelper helps create a dataframe that works with ExcelRewriteUtil
  _rows is a list of dictionaries that can are added to.
  _column_names are the column names for the dataframe.
"""

def generate_col_names(prefix: str) -> str:
    """
    Generator for col00 through col99.
    Copied from PandasUtil.
    Invoke it like this:
      gen = generate_col_names('pfx')
      for i in range(3):
        print next(gen)  # prints pfx00, pfx01, pfx02.

    :return:
    """
    nums = range(100)  # Only 0 .. 99
    for i in nums:
        yield f'{prefix}{i:02d}'


class DfHelper(ExcelUtil):
    def __init__(self):
        super().__init__()
        self._pu = PandasUtil()
        self._du = DateUtil()
        self._cols = None
        self._rows = []
        self._row = {}
        self._static_columns = {}

    @property
    def column_names(self):
        return self._cols

    @property
    def staticColumn(self):
        return self._static_columns
    @staticColumn.setter
    def staticColumn(self, values_dict: dict):
        self._staticColumn = values_dict


    def set_column_names(self, col_name_list: Strings = None, col_count: int = 0):
        """
        Set the column names for the dataframe. Either go from the existing list, or generate names from the col_name
        of the form col00, col01, col02...
        :param col_name_list: list of str with column names. If empty, use col_count. If filled, col_count is ignored.
        :param col_count: how many columns to be used. If col_name_list has one or more names, don't use.
        :return:
        """
        if col_name_list:
            self._cols = col_name_list
        elif col_count:
            gen = generate_col_names('col')
            ans = [next(gen) for k in range(col_count)]
            self._cols = ans
        else:
            self.logger.warning('set_column_names must have either a non-empty list of column names or a column count. No column names set.')

    def build_row(self, col_name: str, value: Union[int, float, str]):
        """
        Add the value to the column name to (but don't save; use add_row_and_clear to do that).
        :param col_name:
        :param value:
        :return:
        """
        self._row[col_name] = value

    def add_static_columns_to_row(self):
        """
        Add static column values (if any) to the current row.
        :return:
        """
        for k, v in self._static_columns.items():
            self.build_row(k, v)

    def add_row_and_clear(self):
        """
        Add the given row and then clear it.
        :return:
        """
        self.add_static_columns_to_row()
        self._rows.append(self._row)
        self._row = {}

    def built_df(self) -> pd.DataFrame:
        """
        Return all the dictionaries as a dataframe.
        :return:
        """
        return self._pu.convert_dict_to_dataframe(self._rows)

    def init_col_value(self, col_name: str, value: Union[int, float, str, datetime]):
        """
        Initialize the col name to the initial value.
        :param col_name: column name to initialize, like 'col00'
        :param value: value to initialize it to.
        :return:
        """
        if not col_name in self.column_names:
            self.logger.warning(f'Unable to find col_name {col_name} in current list. Please initialize with set_column_names.')
            return

        self._static_columns[col_name] = value
        return

    def increment_col_value(self, col_name: str, delta: int = 1, timePeriod: str = None):
        """
        Increment the given col name by delta (which may be negative). If this column contains a numeric, then change it by delta.
        If this column is a datetime, then also look at timePeriod (which should be a string like 'months' or 'days';
        see documentation at https://dateutil.readthedocs.io/en/stable/relativedelta.html.)
        If the column name is a string, then return an error (for now; future versions might go to the next letter in the alphabet).
        :param col_name: column name to initialize, like 'col00'
        :param delta:    a number, like 1
        :param timePeriod: (optional; for datetime only) string that relativeDelta understands, such as 'days', 'weeks', 'months', or 'years'.
        :return:
        """
        if not col_name in self.column_names:
            self.logger.warning(f'Unable to find col_name {col_name} in current list. Taking no action.')
            return
        try:
            cur_value = self._static_columns[col_name]
        except KeyError:
            self.logger.warning(f'Unable to find col_name {col_name} in current list. Please initialize with set_column_names.')
            return
        if isinstance(cur_value, numbers.Number):
            self._static_columns[col_name] = cur_value + delta
        elif isinstance(cur_value, datetime):
            new_date = self._du.changeDate(cur_value, timePeriod=timePeriod, delta=delta)
            self._static_columns[col_name] = new_date
        elif isinstance(cur_value, str):
            self.logger.warning(f'I do not know how to increment {cur_value} yet.')
            raise NotImplementedError
        else:
            self.logger.warning(f'attempting to increment {cur_value}')
            raise TypeError


